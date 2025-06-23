import pandas as pd
from fpdf import FPDF
from datetime import datetime
from collections import defaultdict, OrderedDict
import os

class ReportGenerator:
    """
    Comprehensive report generator for art instruction processing
    Generates reports in Excel, PDF
    """
    
    def __init__(self):
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    def preprocess_report_data(self, report_data):
        """
        Preprocess report data to handle special cases:
        - Convert execution status to NO LOGO for Invalid Logo SKU errors
        - Convert execution status to NOT APPROVED for "Status: Not Approved" errors
        """
        processed_data = []
        for record in report_data:
            processed_record = record.copy()
            error_msg = record.get('Error Message', '')
            
            # Check if error message contains "Invalid Logo SKU:"
            if 'Invalid Logo SKU:' in error_msg and error_msg.strip().endswith('""'):
                processed_record['Execution Status'] = 'NO LOGO'
            # Check if error message is "Status: Not Approved"
            elif error_msg.strip() == "Status: Not Approved":
                processed_record['Execution Status'] = 'NOT APPROVED'
            
            processed_data.append(processed_record)
        
        return processed_data
    
    def create_overview_data(self, report_data):
        """
        Create overview data grouped by document number with completion status
        Preserves the original order from the uploaded file
        """
        # Use OrderedDict to preserve the order of first appearance
        sales_orders = OrderedDict()
        for record in report_data:
            so_number = record.get('Document Number', 'Unknown')
            if so_number not in sales_orders:
                sales_orders[so_number] = []
            sales_orders[so_number].append(record)
        
        overview_data = []
        for so_number, items in sales_orders.items():
            # Calculate counts for this sales order
            so_total = len(items)
            so_success = sum(1 for item in items if item.get('Execution Status') == 'SUCCESS')
            so_failed = sum(1 for item in items if item.get('Execution Status') == 'FAILED')
            so_na = sum(1 for item in items if item.get('Execution Status') == 'N/A')
            so_not_approved = sum(1 for item in items if item.get('Execution Status') == 'Not Approved')
            
            # Calculate success rate (include N/A as success)
            so_success_rate = ((so_success + so_na) / so_total * 100) if so_total > 0 else 0
            
            # Determine completion status based on success rate
            # FULLY SUCCESS: All items are either SUCCESS or N/A (100% success rate)
            # TOTAL FAILED: No items are SUCCESS or N/A (0% success rate) 
            # PARTIAL SUCCESS: Mix of success/failure (1-99% success rate)
            if so_success_rate == 100:
                completion_status = "FULLY SUCCESS"
            elif so_success_rate == 0:
                completion_status = "TOTAL FAILED"
            else:
                completion_status = "PARTIAL SUCCESS"
            
            # Get customer name and due date from first item
            customer_name = items[0].get('Customer/Vendor Name', 'N/A') if items else 'N/A'
            due_date = items[0].get('Due Date', 'N/A') if items else 'N/A'
            
            overview_data.append({
                'Document Number': so_number,
                'Customer/Vendor Name': customer_name,
                'Due Date': due_date,
                'Total Items': so_total,
                'Success': so_success,
                'Failed': so_failed,
                'N/A': so_na,
                'Not Approved': so_not_approved,
                'Success Rate (%)': round(so_success_rate, 1),
                'Completion Status': completion_status
            })
        
        # NO SORTING - preserve original order from OrderedDict
        return overview_data
    
    def calculate_pdf_generation_status(self, items):
        """
        Calculate PDF generation status for a sales order
        Returns a formatted string showing generated PDFs vs total unique logo SKUs
        """
        # Get unique logo SKUs for this sales order (excluding invalid ones)
        unique_logos = set()
        pdf_generated_logos = set()
        
        for item in items:
            logo_sku = str(item.get('LOGO', '')).strip()
            
            # Skip invalid logo SKUs
            if logo_sku and logo_sku not in ['', '0', '0000', 'nan', 'NaN']:
                unique_logos.add(logo_sku)
                
                # Check if PDF was successfully generated for this logo
                execution_status = item.get('Execution Status', '')
                if execution_status == 'SUCCESS':
                    pdf_generated_logos.add(logo_sku)
        
        total_unique_logos = len(unique_logos)
        generated_pdfs = len(pdf_generated_logos)
        
        if total_unique_logos == 0:
            return "0 out of 0 (No valid logos)"
        
        return f"{generated_pdfs} out of {total_unique_logos}"
    
    def generate_all_reports(self, report_data, output_folder, timestamp, sales_order_filter=None):
        """
        Generate all report formats (Excel, PDF)
        
        Args:
            report_data (list): List of dictionaries containing processing results
            output_folder (str): Path to output folder
            timestamp (str): Timestamp for file naming
            sales_order_filter (str): Sales order filter used (if any)
        """
        print(f"Generating comprehensive reports with {len(report_data)} records...")
        
        # Preprocess data to handle Invalid Logo SKU cases
        processed_data = self.preprocess_report_data(report_data)
        
        # Generate each report format
        self.generate_detailed_excel_report(processed_data, output_folder, timestamp, sales_order_filter)
        self.generate_overview_excel_report(processed_data, output_folder, timestamp, sales_order_filter)
        self.generate_pdf_report(processed_data, output_folder, timestamp, sales_order_filter)
        
        print("All reports generated successfully!")
    
    def generate_detailed_excel_report(self, report_data, output_folder, timestamp, sales_order_filter=None):
        """
        Generate detailed Excel report with specified column order and fields
        Preserves the original order from the uploaded file
        """
        if not report_data:
            print("No data to generate detailed Excel report")
            return
            
        try:
            # Convert report data to DataFrame - this preserves the original order
            df = pd.DataFrame(report_data)
            
            # Define the specific columns in the requested order
            detailed_columns = [
                'Document Number',
                'LOGO', 
                'Execution Status',
                'SUBCATEGORY',
                'VENDOR STYLE', 
                'COLOR',
                'SIZE',  # SIZE column as requested
                'Quantity',
                'Customer/Vendor Name',
                'DueDateStatus',
                'Due Date',
                'OPERATIONAL CODE',
                'List of Operation Codes',
                'Error Message'
            ]
            
            # Ensure all required columns exist (add empty columns if missing)
            for col in detailed_columns:
                if col not in df.columns:
                    df[col] = ''
            
            # Format Due Date column to MM/dd/yyyy format
            if 'Due Date' in df.columns:
                df['Due Date'] = df['Due Date'].apply(self.format_date_for_display)
            
            # Format OPERATIONAL CODE column to remove decimal places
            if 'OPERATIONAL CODE' in df.columns:
                df['OPERATIONAL CODE'] = df['OPERATIONAL CODE'].apply(self.format_operational_code)
            
            # Select only the specified columns in the requested order
            df = df[detailed_columns]
            
            # NO SORTING - keep original order from uploaded file
            
            # Generate filename
            filter_suffix = f"_SO_{sales_order_filter}" if sales_order_filter else ""
            filename = f"Art_Instructions_Detailed_Report_{timestamp}{filter_suffix}.xlsx"
            filepath = os.path.join(output_folder, filename)
            
            # Save to Excel with formatting
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Full Detailed Report', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Full Detailed Report']
                
                # Apply formatting
                from openpyxl.styles import PatternFill, Font
                
                # Header formatting
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                
                # Status column formatting
                success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                failed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                no_logo_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Grey for NO LOGO
                not_approved_fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")  # Light orange for NOT APPROVED
                
                execution_status_col = None
                for idx, cell in enumerate(worksheet[1]):
                    if cell.value == 'Execution Status':
                        execution_status_col = idx + 1
                        break
                
                if execution_status_col:
                    for row in range(2, worksheet.max_row + 1):
                        status_cell = worksheet.cell(row=row, column=execution_status_col)
                        if status_cell.value == 'SUCCESS':
                            status_cell.fill = success_fill
                        elif status_cell.value == 'FAILED':
                            status_cell.fill = failed_fill
                        elif status_cell.value == 'NO LOGO':
                            status_cell.fill = no_logo_fill
                        elif status_cell.value == 'NOT APPROVED':
                            status_cell.fill = not_approved_fill
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Max width of 50
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"Detailed Excel report generated: {filename}")
            
        except Exception as e:
            print(f"Error generating detailed Excel report: {e}")
    
    def generate_overview_excel_report(self, report_data, output_folder, timestamp, sales_order_filter=None):
        """
        Generate overview Excel report with Document Number, Completion Status, and PDF Generation Status
        Preserves the original order from the uploaded file
        """
        if not report_data:
            print("No data to generate overview Excel report")
            return
            
        try:
            # Create overview data grouped by document number (preserves original order)
            overview_data = self.create_simple_overview_data(report_data)
            overview_df = pd.DataFrame(overview_data)
            
            # Generate filename
            filter_suffix = f"_SO_{sales_order_filter}" if sales_order_filter else ""
            filename = f"Art_Instructions_Overview_Report_{timestamp}{filter_suffix}.xlsx"
            filepath = os.path.join(output_folder, filename)
            
            # Save to Excel with formatting
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                overview_df.to_excel(writer, sheet_name='Overview Report', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Overview Report']
                
                # Apply formatting
                from openpyxl.styles import PatternFill, Font
                
                # Header formatting
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                
                # Status column formatting
                fully_success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                partial_success_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")  # Dark Blue
                partial_success_font = Font(color="FFFFFF")  # White text for dark blue background
                total_failed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
                not_approved_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange for NOT APPROVED
                not_approved_font = Font(color="FFFFFF")  # White text for orange background
                
                # Find completion status column
                completion_status_col = None
                for idx, cell in enumerate(worksheet[1]):
                    if cell.value == 'Completion Status':
                        completion_status_col = idx + 1
                        break
                
                if completion_status_col:
                    for row in range(2, worksheet.max_row + 1):
                        status_cell = worksheet.cell(row=row, column=completion_status_col)
                        if status_cell.value == 'FULLY SUCCESS':
                            status_cell.fill = fully_success_fill
                        elif status_cell.value == 'PARTIAL SUCCESS':
                            status_cell.fill = partial_success_fill
                            status_cell.font = partial_success_font
                        elif status_cell.value == 'TOTAL FAILED':
                            status_cell.fill = total_failed_fill
                        elif status_cell.value == 'NOT APPROVED':
                            status_cell.fill = not_approved_fill
                            status_cell.font = not_approved_font
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)  # Max width of 30 for overview
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"Overview Excel report generated: {filename}")
            
        except Exception as e:
            print(f"Error generating overview Excel report: {e}")
    
    def create_simple_overview_data(self, report_data):
        """
        Create simple overview data with Document Number, Completion Status, and PDF Generation Status
        Preserves the original order from the uploaded file
        """
        # Use OrderedDict to preserve the order of first appearance
        sales_orders = OrderedDict()
        for record in report_data:
            so_number = record.get('Document Number', 'Unknown')
            if so_number not in sales_orders:
                sales_orders[so_number] = []
            sales_orders[so_number].append(record)
        
        overview_data = []
        for so_number, items in sales_orders.items():
            # Calculate counts for this sales order
            so_total = len(items)
            so_success = sum(1 for item in items if item.get('Execution Status') == 'SUCCESS')
            so_no_logo = sum(1 for item in items if item.get('Execution Status') == 'NO LOGO')
            so_failed = sum(1 for item in items if item.get('Execution Status') == 'FAILED')
            so_not_approved = sum(1 for item in items if item.get('Execution Status') == 'NOT APPROVED')
            
            # Calculate success rate (include only NO LOGO as success, NOT APPROVED is considered failure)
            so_success_rate = ((so_success + so_no_logo) / so_total * 100) if so_total > 0 else 0
            
            # Determine completion status with special handling for NOT APPROVED
            # NOT APPROVED: If ALL items are NOT APPROVED (no SUCCESS, FAILED, or NO LOGO items)
            # FULLY SUCCESS: All items are either SUCCESS or NO LOGO (100% success rate)
            # TOTAL FAILED: No items are SUCCESS or NO LOGO (0% success rate) 
            # PARTIAL SUCCESS: Mix of success/failure (1-99% success rate)
            if so_not_approved == so_total:  # All items are NOT APPROVED
                completion_status = "NOT APPROVED"
            elif so_success_rate == 100:
                completion_status = "FULLY SUCCESS"
            elif so_success_rate == 0:
                completion_status = "TOTAL FAILED"
            else:
                completion_status = "PARTIAL SUCCESS"
            
            # Calculate PDF generation status
            pdf_generation_status = self.calculate_pdf_generation_status(items)
            
            overview_data.append({
                'Document Number': so_number,
                'Completion Status': completion_status,
                'PDF Generation Status': pdf_generation_status
            })
        
        # NO SORTING - preserve original order from OrderedDict
        return overview_data
    
    def generate_pdf_report(self, report_data, output_folder, timestamp, sales_order_filter=None):
        """
        Generate PDF report organized by sales order with item-level details
        Preserves the original order from the uploaded file
        """
        if not report_data:
            print("No data to generate PDF report")
            return
            
        try:
            # Use OrderedDict to preserve the order of first appearance
            sales_orders = OrderedDict()
            for record in report_data:
                so_number = record.get('Document Number', 'Unknown')
                if so_number not in sales_orders:
                    sales_orders[so_number] = []
                sales_orders[so_number].append(record)
            
            # Generate filename
            filter_suffix = f"_SO_{sales_order_filter}" if sales_order_filter else ""
            filename = f"Art_Instructions_Report_{timestamp}{filter_suffix}.pdf"
            filepath = os.path.join(output_folder, filename)
            
            # Create PDF
            pdf = FPDF()
            pdf.set_auto_page_break(auto=True, margin=15)
            
            # Title page
            pdf.add_page()
            pdf.set_font('Arial', 'B', 16)
            pdf.cell(0, 10, 'Art Instructions Processing Report', ln=True, align='C')
            pdf.ln(5)
            
            pdf.set_font('Arial', '', 12)
            pdf.cell(0, 8, f'Generated: {datetime.now().strftime("%m/%d/%Y %H:%M:%S")}', ln=True, align='C')
            
            if sales_order_filter:
                pdf.cell(0, 8, f'Filtered by Sales Order: {sales_order_filter}', ln=True, align='C')
            
            pdf.ln(10)
            
            # Summary statistics (removed Total Sales Orders)
            total_records = len(report_data)
            success_count = sum(1 for record in report_data if record.get('Execution Status') == 'SUCCESS')
            failed_count = sum(1 for record in report_data if record.get('Execution Status') == 'FAILED')
            no_logo_count = sum(1 for record in report_data if record.get('Execution Status') == 'NO LOGO')
            not_approved_count = sum(1 for record in report_data if record.get('Execution Status') == 'NOT APPROVED')
            # Include only NO LOGO as success for success rate calculation (NOT APPROVED is considered failure)
            success_rate = ((success_count + no_logo_count) / total_records * 100) if total_records > 0 else 0
            
            pdf.set_font('Arial', 'B', 14)
            pdf.cell(0, 8, 'Summary Statistics', ln=True)
            pdf.ln(2)
            
            pdf.set_font('Arial', '', 10)
            pdf.cell(0, 6, f'Total Records Processed: {total_records}', ln=True)
            pdf.cell(0, 6, f'Successful: {success_count}', ln=True)
            pdf.cell(0, 6, f'Failed: {failed_count}', ln=True)
            pdf.cell(0, 6, f'NO LOGO (Invalid Logo SKU): {no_logo_count}', ln=True)
            pdf.cell(0, 6, f'NOT APPROVED: {not_approved_count}', ln=True)
            pdf.cell(0, 6, f'Success Rate: {success_rate:.1f}% (includes NO LOGO as success, NOT APPROVED as failure)', ln=True)
            
            pdf.ln(10)
            
            # Sales Orders Summary Statistics
            so_fully_success = 0
            so_partial_success = 0
            so_not_approved = 0
            so_total_failed = 0
            
            for so_number, items in sales_orders.items():
                so_total = len(items)
                so_success = sum(1 for item in items if item.get('Execution Status') == 'SUCCESS')
                so_no_logo = sum(1 for item in items if item.get('Execution Status') == 'NO LOGO')
                so_failed = sum(1 for item in items if item.get('Execution Status') == 'FAILED')
                so_not_approved_items = sum(1 for item in items if item.get('Execution Status') == 'NOT APPROVED')
                
                # Calculate success rate for this SO
                so_success_rate = ((so_success + so_no_logo) / so_total * 100) if so_total > 0 else 0
                
                # Categorize this sales order
                if so_not_approved_items == so_total:  # All items are NOT APPROVED
                    so_not_approved += 1
                elif so_success_rate == 100:
                    so_fully_success += 1
                elif so_success_rate == 0:
                    so_total_failed += 1
                else:
                    so_partial_success += 1
            
            pdf.set_font('Arial', 'B', 14)
            pdf.cell(0, 8, 'Sales Orders Summary Statistics', ln=True)
            pdf.ln(2)
            
            # Calculate sales orders success rate
            so_success_rate = (so_fully_success / len(sales_orders) * 100) if len(sales_orders) > 0 else 0
            
            pdf.set_font('Arial', '', 10)
            pdf.cell(0, 6, f'Total Sales Orders: {len(sales_orders)}', ln=True)
            pdf.cell(0, 6, f'No of Sales Orders Fully Success: {so_fully_success}', ln=True)
            pdf.cell(0, 6, f'No of Sales Orders Partial Success: {so_partial_success}', ln=True)
            pdf.cell(0, 6, f'No of Sales Orders Total Failed: {so_total_failed}', ln=True)
            pdf.cell(0, 6, f'No of Sales Orders Not Approved: {so_not_approved}', ln=True)
            pdf.cell(0, 6, f'Sales Orders Success Rate: {so_success_rate:.1f}% ({so_fully_success} out of {len(sales_orders)})', ln=True)
            
            pdf.ln(10)
            
            # Detailed report by sales order (preserves original order)
            pdf.set_font('Arial', 'B', 14)
            pdf.cell(0, 8, 'Detailed Report by Sales Order', ln=True)
            pdf.ln(5)
            
            for so_number, items in sales_orders.items():
                # Check if we need a new page
                if pdf.get_y() > 250:
                    pdf.add_page()
                
                # Sales Order header
                pdf.set_font('Arial', 'B', 12)
                pdf.cell(0, 8, f'Sales Order: {so_number}', ln=True)
                pdf.ln(2)
                
                # SO summary (updated to include NO LOGO, NOT APPROVED counts and success rate)
                so_total = len(items)
                so_success = sum(1 for item in items if item.get('Execution Status') == 'SUCCESS')
                so_failed = sum(1 for item in items if item.get('Execution Status') == 'FAILED')
                so_no_logo = sum(1 for item in items if item.get('Execution Status') == 'NO LOGO')
                so_not_approved = sum(1 for item in items if item.get('Execution Status') == 'NOT APPROVED')
                # Include only NO LOGO as success for success rate calculation (NOT APPROVED is considered failure)
                so_success_rate = ((so_success + so_no_logo) / so_total * 100) if so_total > 0 else 0
                
                # Calculate PDF generation status
                pdf_generation_status = self.calculate_pdf_generation_status(items)
                
                # Determine completion status with special handling for NOT APPROVED
                # NOT APPROVED: If ALL items are NOT APPROVED (no SUCCESS, FAILED, or NO LOGO items)
                # FULLY SUCCESS: All items are either SUCCESS or NO LOGO (100% success rate)
                # TOTAL FAILED: No items are SUCCESS or NO LOGO (0% success rate) 
                # PARTIAL SUCCESS: Mix of success/failure (1-99% success rate)
                if so_not_approved == so_total:  # All items are NOT APPROVED
                    completion_status = "NOT APPROVED"
                    status_color = (255, 165, 0)  # Orange
                elif so_success_rate == 100:
                    completion_status = "FULLY SUCCESS"
                    status_color = (0, 128, 0)  # Green
                elif so_success_rate == 0:
                    completion_status = "TOTAL FAILED"
                    status_color = (255, 0, 0)  # Red
                else:
                    completion_status = "PARTIAL SUCCESS"
                    status_color = (0, 0, 139)  # Dark Blue
                
                pdf.set_font('Arial', '', 10)
                pdf.cell(0, 5, f'Items: {so_total} | Success: {so_success} | Failed: {so_failed} | NO LOGO: {so_no_logo} | NOT APPROVED: {so_not_approved}', ln=True)
                pdf.cell(0, 5, f'Success Rate: {so_success_rate:.1f}% (includes NO LOGO as success, NOT APPROVED as failure)', ln=True)
                pdf.cell(0, 5, f'PDF Generation Status: {pdf_generation_status}', ln=True)
                
                # Add completion status with color
                pdf.set_text_color(status_color[0], status_color[1], status_color[2])
                pdf.set_font('Arial', 'B', 10)
                pdf.cell(0, 5, f'Completion Status: {completion_status}', ln=True)
                pdf.set_text_color(0, 0, 0)  # Reset to black
                pdf.ln(3)
                
                # Customer info (from first item)
                if items:
                    customer_name = items[0].get('Customer/Vendor Name', 'N/A')
                    due_date_raw = items[0].get('Due Date', 'N/A')
                    # Format due date
                    due_date = self.format_date_for_display(due_date_raw) if due_date_raw != 'N/A' else 'N/A'
                    pdf.cell(0, 5, f'Customer: {customer_name}', ln=True)
                    pdf.cell(0, 5, f'Due Date: {due_date}', ln=True)
                    pdf.ln(3)
                
                # Items table header with new column order and widths
                pdf.set_font('Arial', 'B', 8)
                # Column widths: Logo(20), Status(20), Error Message(35), Description(25), Style(20), Color(15), Size(10), Qty(10), Op Code(10)
                col_widths = [20, 20, 35, 25, 20, 15, 10, 10, 10]
                headers = ['Logo', 'Status', 'Error Message', 'Description', 'Style', 'Color', 'Size', 'Qty', 'Op Code']
                
                for i, header in enumerate(headers):
                    pdf.cell(col_widths[i], 6, header, 1, 0, 'C')
                pdf.ln()
                
                # Items data with multi-line support
                pdf.set_font('Arial', '', 7)
                for item in items:
                    # Check if we need a new page
                    if pdf.get_y() > 260:
                        pdf.add_page()
                        # Repeat header on new page
                        pdf.set_font('Arial', 'B', 8)
                        for i, header in enumerate(headers):
                            pdf.cell(col_widths[i], 6, header, 1, 0, 'C')
                        pdf.ln()
                        pdf.set_font('Arial', '', 7)
                    
                    # Prepare values with formatting
                    logo_val = str(item.get('LOGO', ''))
                    status_val = str(item.get('Execution Status', ''))
                    error_val = str(item.get('Error Message', ''))
                    desc_val = str(item.get('SUBCATEGORY', ''))
                    style_val = str(item.get('VENDOR STYLE', ''))
                    color_val = str(item.get('COLOR', ''))
                    size_val = str(item.get('SIZE', ''))
                    qty_val = str(item.get('Quantity', ''))
                    op_code_raw = item.get('OPERATIONAL CODE', '')
                    op_code_val = self.format_operational_code(op_code_raw)
                    
                    values = [logo_val, status_val, error_val, desc_val, style_val, color_val, size_val, qty_val, op_code_val]
                    
                    # Calculate required height for this row based on text wrapping
                    row_height = self.calculate_row_height(pdf, values, col_widths)
                    
                    # Store current position
                    start_x = pdf.get_x()
                    start_y = pdf.get_y()
                    
                    # Draw cells with proper text wrapping
                    for i, (value, width) in enumerate(zip(values, col_widths)):
                        cell_x = start_x + sum(col_widths[:i])
                        
                        # Draw cell border
                        pdf.rect(cell_x, start_y, width, row_height)
                        
                        # Add text with wrapping
                        self.add_wrapped_text(pdf, value, cell_x, start_y, width, row_height)
                    
                    # Move to next row
                    pdf.set_xy(start_x, start_y + row_height)
                
                pdf.ln(5)  # Space between sales orders
            
            pdf.output(filepath)
            print(f"PDF report generated: {filename}")
            
        except Exception as e:
            print(f"Error generating PDF report: {e}")

    
    def get_error_statistics(self, report_data):
        """
        Get detailed error statistics for debugging purposes
        Updated to handle NO LOGO and Not Approved status separately
        """
        error_stats = {
            "total_errors": 0,
            "total_no_logo": 0,
            "total_not_approved": 0,
            "error_types": defaultdict(int),
            "errors_by_sales_order": defaultdict(list),
            "no_logo_by_sales_order": defaultdict(list),
            "not_approved_by_sales_order": defaultdict(list),
            "most_common_errors": []
        }
        
        for record in report_data:
            status = record.get('Execution Status')
            so_number = record.get('Document Number', 'Unknown')
            
            if status == 'FAILED':
                error_stats["total_errors"] += 1
                error_msg = record.get('Error Message', 'Unknown error')
                error_stats["error_types"][error_msg] += 1
                
                error_stats["errors_by_sales_order"][so_number].append({
                    "logo": record.get('LOGO', 'N/A'),
                    "error": error_msg,
                    "style": record.get('VENDOR STYLE', 'N/A')
                })
            elif status == 'NO LOGO':
                error_stats["total_no_logo"] += 1
                error_stats["no_logo_by_sales_order"][so_number].append({
                    "logo": record.get('LOGO', 'N/A'),
                    "error": record.get('Error Message', 'Invalid Logo SKU'),
                    "style": record.get('VENDOR STYLE', 'N/A')
                })
            elif status == 'Not Approved':
                error_stats["total_not_approved"] += 1
                error_stats["not_approved_by_sales_order"][so_number].append({
                    "logo": record.get('LOGO', 'N/A'),
                    "error": record.get('Error Message', 'Status: Not Approved'),
                    "style": record.get('VENDOR STYLE', 'N/A')
                })
        
        # Sort errors by frequency
        error_stats["most_common_errors"] = sorted(
            error_stats["error_types"].items(), 
            key=lambda x: x[1], 
            reverse=True
        )
        
        return dict(error_stats)
    
    def format_date_for_display(self, date_value):
        """
        Format date values to MM/dd/yyyy format for display in reports
        """
        if pd.isna(date_value) or date_value == "" or str(date_value).strip() == "":
            return ""
        
        try:
            # Handle different input types
            if isinstance(date_value, str):
                date_str = str(date_value).strip()
                
                # If it's already in MM/dd/yyyy format, return as-is
                if '/' in date_str and len(date_str.split('/')) == 3:
                    parts = date_str.split('/')
                    if len(parts[2]) == 4:  # Already in MM/dd/yyyy format
                        return date_str
                
                # Try to parse various string formats
                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%Y/%m/%d', '%d/%m/%Y']:
                    try:
                        date_obj = datetime.strptime(date_str, fmt)
                        return date_obj.strftime('%m/%d/%Y')
                    except ValueError:
                        continue
                
                # If no format worked, return original
                return date_str
            
            elif isinstance(date_value, (int, float)):
                # Excel serial date number
                if date_value > 25000:  # Reasonable range for Excel dates
                    excel_epoch = datetime(1899, 12, 30)
                    date_obj = excel_epoch + pd.Timedelta(days=date_value)
                    return date_obj.strftime('%m/%d/%Y')
                else:
                    return str(int(date_value))
            else:
                # Pandas datetime or other datetime object
                date_obj = pd.to_datetime(date_value)
                return date_obj.strftime('%m/%d/%Y')
                
        except Exception as e:
            print(f"Error formatting date '{date_value}': {e}")
            return str(date_value)
    
    def format_operational_code(self, op_code_value):
        """
        Format operational code to remove decimal places (11.0 -> 11)
        """
        if pd.isna(op_code_value) or op_code_value == "" or str(op_code_value).strip() == "":
            return ""
        
        try:
            # Convert to string and check if it's a number
            op_code_str = str(op_code_value).strip()
            
            # If it's a float-like number (e.g., "11.0"), convert to integer
            if '.' in op_code_str and op_code_str.replace('.', '').isdigit():
                try:
                    float_val = float(op_code_str)
                    if float_val.is_integer():
                        return str(int(float_val))
                    else:
                        return op_code_str
                except ValueError:
                    return op_code_str
            
            # If it's already an integer or doesn't have decimal, return as-is
            return op_code_str
            
        except Exception as e:
            print(f"Error formatting operational code '{op_code_value}': {e}")
            return str(op_code_value)
    
    def calculate_row_height(self, pdf, values, col_widths):
        """
        Calculate the required height for a table row based on text wrapping
        """
        max_lines = 1
        
        for value, width in zip(values, col_widths):
            if not value:
                continue
                
            # Calculate available width (subtract padding)
            available_width = width - 2
            
            # Get text width
            text_width = pdf.get_string_width(str(value))
            
            # Calculate number of lines needed
            if text_width > available_width:
                lines_needed = int(text_width / available_width) + 1
                max_lines = max(max_lines, lines_needed)
        
        # Return height (base height * number of lines)
        return max_lines * 5
    
    def add_wrapped_text(self, pdf, text, x, y, width, height):
        """
        Add text to a cell with proper wrapping
        """
        if not text:
            return
        
        text = str(text)
        available_width = width - 2  # Subtract padding
        line_height = 5
        max_lines = int(height / line_height)
        
        # Set position with padding
        pdf.set_xy(x + 1, y + 1)
        
        # If text fits in one line
        if pdf.get_string_width(text) <= available_width:
            pdf.cell(width - 2, line_height, text, 0, 0, 'L')
            return
        
        # Split text into words
        words = text.split()
        lines = []
        current_line = ""
        
        for word in words:
            test_line = current_line + (" " if current_line else "") + word
            if pdf.get_string_width(test_line) <= available_width:
                current_line = test_line
            else:
                if current_line:
                    lines.append(current_line)
                    current_line = word
                else:
                    # Single word is too long, truncate it
                    truncated_word = word
                    while pdf.get_string_width(truncated_word + "...") > available_width and len(truncated_word) > 1:
                        truncated_word = truncated_word[:-1]
                    lines.append(truncated_word + "...")
                    current_line = ""
        
        if current_line:
            lines.append(current_line)
        
        # Limit to max lines that fit in cell height
        lines = lines[:max_lines]
        
        # Add lines to PDF
        for i, line in enumerate(lines):
            pdf.set_xy(x + 1, y + 1 + (i * line_height))
            pdf.cell(width - 2, line_height, line, 0, 0, 'L')